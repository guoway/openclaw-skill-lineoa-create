"""Python 版最小可用 indexer 服務。"""

from __future__ import annotations

import asyncio
import hashlib
import os
import uuid
from dataclasses import dataclass
from pathlib import Path

import docx
import httpx
import xlrd
from openpyxl import load_workbook
from pypdf import PdfReader
from sqlalchemy.orm import Session
from watchdog.events import FileSystemEvent, FileSystemEventHandler
from watchdog.observers import Observer

from app.core.config import Settings
from app.core.logging import format_log_context, get_logger
from app.repositories.documents import DocumentRepository
from app.services.rag_service import RagService


logger = get_logger(__name__)

SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".txt", ".md", ".markdown", ".xlsx", ".xls"}


@dataclass
class IndexingResult:
    """單一文件索引結果。"""

    file_path: str
    status: str
    chunk_count: int


class IndexerService:
    """提供最小可用的文件掃描與索引能力。"""

    def __init__(self, settings: Settings, db: Session) -> None:
        self.settings = settings
        self.db = db
        self.document_repository = DocumentRepository(db)
        self.rag_service = RagService(settings)

    async def index_all(self) -> list[IndexingResult]:
        """掃描 knowledge 目錄並索引所有支援檔案。"""

        knowledge_dir = Path(self.settings.knowledge_base_dir)
        if not knowledge_dir.exists():
            return []

        results: list[IndexingResult] = []
        for path in sorted(knowledge_dir.rglob("*")):
            if not path.is_file() or path.suffix.lower() not in SUPPORTED_EXTENSIONS:
                continue
            results.append(await self.index_file(path))
        self.db.commit()
        return results

    async def delete_file(self, path: str | Path) -> IndexingResult:
        """刪除文件對應的 Qdrant 與 t_documents 紀錄。"""

        file_path = Path(path)
        doc_id = self._build_doc_id(file_path)
        await self._delete_old_chunks(doc_id)
        deleted = self.document_repository.delete_by_doc_id(doc_id)
        self.db.commit()
        return IndexingResult(file_path=str(file_path), status="deleted" if deleted else "not_found", chunk_count=0)

    async def watch(self) -> None:
        """長駐監看 knowledge 目錄變化。"""

        knowledge_dir = Path(self.settings.knowledge_base_dir)
        if not knowledge_dir.exists():
            return

        loop = asyncio.get_running_loop()
        event_handler = _IndexerWatchHandler(service=self, loop=loop)
        observer = Observer()
        observer.schedule(event_handler, str(knowledge_dir), recursive=True)
        observer.start()
        try:
            while True:
                await asyncio.sleep(1)
        finally:
            observer.stop()
            observer.join()

    async def index_file(self, path: str | Path) -> IndexingResult:
        """索引單一文件。"""

        file_path = Path(path)
        file_hash = self._get_file_hash(file_path)
        doc_id = self._build_doc_id(file_path)
        existing = self.document_repository.get_by_doc_id(doc_id)

        if existing and existing.content_hash == file_hash and existing.status == "indexed":
            logger.info("index skipped | %s", format_log_context(file_path=str(file_path), status="skipped"))
            return IndexingResult(file_path=str(file_path), status="skipped", chunk_count=existing.chunk_count)

        self.document_repository.upsert_processing_record(
            doc_id=doc_id,
            filename=file_path.name,
            file_path=str(file_path),
            file_type=file_path.suffix.lower().lstrip("."),
            file_size=file_path.stat().st_size,
            content_hash=file_hash,
            status="processing",
        )

        try:
            content = self._parse_file(file_path)
            chunks = self._split_into_chunks(content)
            await self._delete_old_chunks(doc_id)
            await self._upload_chunks(chunks, file_path=file_path, source_id=doc_id, content_hash=file_hash)
            self.document_repository.upsert_processing_record(
                doc_id=doc_id,
                filename=file_path.name,
                file_path=str(file_path),
                file_type=file_path.suffix.lower().lstrip("."),
                file_size=file_path.stat().st_size,
                content_hash=file_hash,
                status="indexed",
                chunk_count=len(chunks),
                indexed=True,
            )
            self.db.flush()
            logger.info(
                "index completed | %s",
                format_log_context(file_path=str(file_path), chunk_count=len(chunks), status="indexed"),
            )
            return IndexingResult(file_path=str(file_path), status="indexed", chunk_count=len(chunks))
        except Exception as exc:
            self.document_repository.upsert_processing_record(
                doc_id=doc_id,
                filename=file_path.name,
                file_path=str(file_path),
                file_type=file_path.suffix.lower().lstrip("."),
                file_size=file_path.stat().st_size,
                content_hash=file_hash,
                status="failed",
                chunk_count=0,
                error_message=str(exc),
            )
            self.db.flush()
            logger.exception(
                "index failed | %s",
                format_log_context(file_path=str(file_path), error=str(exc), status="failed"),
            )
            return IndexingResult(file_path=str(file_path), status="failed", chunk_count=0)

    def _build_doc_id(self, file_path: Path) -> str:
        """依檔案路徑建立穩定 doc_id。"""

        return hashlib.md5(str(file_path).encode("utf-8")).hexdigest()

    def _get_file_hash(self, file_path: Path) -> str:
        """取得檔案 sha256。"""

        return hashlib.sha256(file_path.read_bytes()).hexdigest()

    def _parse_file(self, file_path: Path) -> str:
        """依副檔名解析文件內容。"""

        suffix = file_path.suffix.lower()
        if suffix == ".pdf":
            return self._parse_pdf(file_path)
        if suffix == ".docx":
            return self._parse_docx(file_path)
        if suffix in {".txt", ".md", ".markdown"}:
            return self._read_text_with_fallback(file_path)
        if suffix in {".xlsx", ".xls"}:
            return self._parse_excel(file_path)
        raise ValueError(f"Unsupported file type: {suffix}")

    def _parse_pdf(self, file_path: Path) -> str:
        reader = PdfReader(str(file_path))
        return "\n".join(page.extract_text() or "" for page in reader.pages)

    def _parse_docx(self, file_path: Path) -> str:
        document = docx.Document(str(file_path))
        return "\n".join(paragraph.text for paragraph in document.paragraphs)

    def _parse_excel(self, file_path: Path) -> str:
        if file_path.suffix.lower() == ".xls":
            return self._parse_xls(file_path)

        workbook = load_workbook(filename=str(file_path), data_only=True)
        all_lines: list[str] = []
        for sheet in workbook.worksheets:
            rows: list[str] = []
            for row in sheet.iter_rows(values_only=True):
                values = [str(cell).strip() for cell in row if cell not in (None, "")]
                if values:
                    rows.append(", ".join(values))
            if rows:
                all_lines.append(f"=== 工作表: {sheet.title} ===")
                all_lines.extend(rows)
        return "\n".join(all_lines)

    def _parse_xls(self, file_path: Path) -> str:
        """解析舊版 .xls。"""

        workbook = xlrd.open_workbook(str(file_path))
        all_lines: list[str] = []
        for sheet in workbook.sheets():
            rows: list[str] = []
            for row_index in range(sheet.nrows):
                values = [str(cell).strip() for cell in sheet.row_values(row_index) if cell not in (None, "")]
                if values:
                    rows.append(", ".join(values))
            if rows:
                all_lines.append(f"=== 工作表: {sheet.name} ===")
                all_lines.extend(rows)
        return "\n".join(all_lines)

    def _read_text_with_fallback(self, file_path: Path) -> str:
        """以多種常見編碼讀取文字檔。"""

        for encoding in ("utf-8", "utf-8-sig", "cp950", "big5", "latin-1"):
            try:
                return file_path.read_text(encoding=encoding)
            except UnicodeDecodeError:
                continue
        raise UnicodeDecodeError("text", b"", 0, 1, f"Unsupported text encoding: {file_path}")

    def _split_into_chunks(self, text: str, *, chunk_size: int = 1000, overlap: int = 200) -> list[str]:
        """切 chunk 並略過過短內容。"""

        chunks: list[str] = []
        index = 0
        while index < len(text):
            chunk = text[index : index + chunk_size]
            content = chunk.replace(",", "").replace(" ", "").replace("\n", "").replace("\r", "").replace("\t", "")
            if len(content) > 10:
                chunks.append(chunk)
            index += chunk_size - overlap
        return chunks

    async def _delete_old_chunks(self, source_id: str) -> None:
        """刪除舊的 Qdrant chunks。"""

        async with httpx.AsyncClient(timeout=30) as client:
            response = await client.post(
                f"{self.settings.qdrant_url}/collections/{self.settings.qdrant_collection}/points/delete",
                json={
                    "filter": {
                        "must": [
                            {"key": "metadata.source_id", "match": {"value": source_id}}
                        ]
                    }
                },
            )
            response.raise_for_status()

    async def _upload_chunks(self, chunks: list[str], *, file_path: Path, source_id: str, content_hash: str) -> None:
        """上傳 chunks 到 Qdrant。"""

        points = []
        for index, chunk in enumerate(chunks):
            vector = await self.rag_service.get_embedding(chunk)
            points.append(
                {
                    "id": str(uuid.uuid4()),
                    "vector": vector,
                    "payload": {
                        "content": chunk,
                        "source": file_path.name,
                        "metadata": {
                            "source_id": source_id,
                            "filename": file_path.name,
                            "filepath": str(file_path),
                            "content_hash": content_hash,
                            "chunk_index": index,
                            "total_chunks": len(chunks),
                        },
                    },
                }
            )

        batch_size = 50
        async with httpx.AsyncClient(timeout=30) as client:
            for offset in range(0, len(points), batch_size):
                batch = points[offset : offset + batch_size]
                response = await client.put(
                    f"{self.settings.qdrant_url}/collections/{self.settings.qdrant_collection}/points",
                    json={"points": batch},
                )
                response.raise_for_status()


class _IndexerWatchHandler(FileSystemEventHandler):
    """將檔案系統事件轉交給 IndexerService。"""

    def __init__(self, *, service: IndexerService, loop: asyncio.AbstractEventLoop) -> None:
        self.service = service
        self.loop = loop

    def on_created(self, event: FileSystemEvent) -> None:
        if not event.is_directory:
            self._submit_index(event.src_path)

    def on_modified(self, event: FileSystemEvent) -> None:
        if not event.is_directory:
            self._submit_index(event.src_path)

    def on_deleted(self, event: FileSystemEvent) -> None:
        if not event.is_directory:
            future = asyncio.run_coroutine_threadsafe(self.service.delete_file(event.src_path), self.loop)
            future.add_done_callback(self._log_future_exception)

    def _submit_index(self, file_path: str) -> None:
        if Path(file_path).suffix.lower() not in SUPPORTED_EXTENSIONS:
            return
        future = asyncio.run_coroutine_threadsafe(self.service.index_file(file_path), self.loop)
        future.add_done_callback(self._log_future_exception)

    @staticmethod
    def _log_future_exception(future) -> None:
        try:
            future.result()
        except Exception as exc:
            logger.exception("indexer watch task failed | %s", format_log_context(error=str(exc)))
